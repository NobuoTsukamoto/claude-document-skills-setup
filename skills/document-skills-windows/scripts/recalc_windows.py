# -*- coding: utf-8 -*-
"""LibreOffice で xlsx の全数式を再計算して保存する (Windows 用)。

公式 xlsx スキルの scripts/recalc.py は Unix 専用（soffice.py の AF_UNIX 依存・
マクロ配置先が macOS/Linux 固定）のため、Windows ではこのスクリプトを使う。

やること:
  1. LibreOffice ユーザープロファイル (%APPDATA%\\LibreOffice) に再計算マクロを配置
  2. soffice --headless でマクロを実行し、全数式を再計算して上書き保存
  3. openpyxl (data_only=True) で Excel エラー値と数式数を確認し JSON で報告
     （公式 recalc.py と同じ出力形式: status / total_errors / total_formulas / error_summary）

Usage:
    python recalc_windows.py <file.xlsx> [timeout_seconds]
"""
import json
import os
import subprocess
import sys
from pathlib import Path

from openpyxl import load_workbook

MACRO_NAME = "RecalcAndStore"

# LibreOffice Basic の標準的な再計算イディオム（calculateAll → store → close）
MODULE_XML = f"""<?xml version="1.0" encoding="UTF-8"?>
<!DOCTYPE script:module PUBLIC "-//OpenOffice.org//DTD OfficeDocument 1.0//EN" "module.dtd">
<script:module xmlns:script="http://openoffice.org/2000/script" script:name="Module1" script:language="StarBasic">
    Sub {MACRO_NAME}()
      ThisComponent.calculateAll()
      ThisComponent.store()
      ThisComponent.close(True)
    End Sub
</script:module>
"""

SCRIPT_XLB = """<?xml version="1.0" encoding="UTF-8"?>
<!DOCTYPE library:library PUBLIC "-//OpenOffice.org//DTD OfficeDocument 1.0//EN" "library.dtd">
<library:library xmlns:library="http://openoffice.org/2000/library" library:name="Standard" library:readonly="false" library:passwordprotected="false">
 <library:element library:name="Module1"/>
</library:library>
"""

EXCEL_ERRORS = ("#VALUE!", "#DIV/0!", "#REF!", "#NAME?", "#NULL!", "#NUM!", "#N/A")


def ensure_macro():
    """再計算マクロを LibreOffice ユーザープロファイルに配置する（冪等）。"""
    std = Path(os.environ["APPDATA"]) / "LibreOffice" / "4" / "user" / "basic" / "Standard"
    if not std.exists():
        # 初回起動でユーザープロファイルを生成させる
        subprocess.run(["soffice", "--headless", "--terminate_after_init"],
                       capture_output=True, timeout=60)
        std.mkdir(parents=True, exist_ok=True)

    xlb = std / "script.xlb"
    if xlb.exists():
        text = xlb.read_text(encoding="utf-8")
        if "Module1" not in text:
            text = text.replace(
                "</library:library>",
                ' <library:element library:name="Module1"/>\n</library:library>')
            xlb.write_text(text, encoding="utf-8")
    else:
        xlb.write_text(SCRIPT_XLB, encoding="utf-8")

    (std / "Module1.xba").write_text(MODULE_XML, encoding="utf-8")


def verify(path: Path):
    """再計算後のブックを openpyxl で検証する。"""
    wb = load_workbook(path, data_only=True)
    error_locations = {}
    total_errors = 0
    for sheet in wb.sheetnames:
        for row in wb[sheet].iter_rows():
            for cell in row:
                if isinstance(cell.value, str):
                    for err in EXCEL_ERRORS:
                        if err in cell.value:
                            error_locations.setdefault(err, []).append(
                                f"{sheet}!{cell.coordinate}")
                            total_errors += 1
                            break
    wb.close()

    wb = load_workbook(path, data_only=False)
    total_formulas = sum(
        1 for sheet in wb.sheetnames for row in wb[sheet].iter_rows()
        for cell in row
        if isinstance(cell.value, str) and cell.value.startswith("="))
    wb.close()

    return {
        "status": "success" if total_errors == 0 else "errors_found",
        "total_errors": total_errors,
        "total_formulas": total_formulas,
        "error_summary": {
            err: {"count": len(locs), "locations": locs[:20]}
            for err, locs in error_locations.items()
        },
    }


def recalc(path: Path, timeout: int):
    if not path.exists():
        return {"error": f"File {path} does not exist"}

    ensure_macro()

    macro_url = (f"vnd.sun.star.script:Standard.Module1.{MACRO_NAME}"
                 "?language=Basic&location=application")
    try:
        proc = subprocess.run(
            ["soffice", "--headless", "--norestore", macro_url, str(path.resolve())],
            capture_output=True, text=True, timeout=timeout)
    except FileNotFoundError:
        return {"error": "soffice not found on PATH (run setup-document-skills.ps1)"}
    except subprocess.TimeoutExpired:
        return {"error": f"soffice timed out after {timeout}s "
                         "(close any running LibreOffice windows and retry)"}
    if proc.returncode != 0:
        return {"error": proc.stderr.strip() or f"soffice exited with {proc.returncode}"}

    try:
        return verify(path)
    except Exception as e:  # noqa: BLE001
        return {"error": str(e)}


def main():
    if len(sys.argv) < 2:
        print(__doc__)
        sys.exit(1)
    path = Path(sys.argv[1])
    timeout = int(sys.argv[2]) if len(sys.argv) > 2 else 60
    result = recalc(path, timeout)
    print(json.dumps(result, indent=2, ensure_ascii=False))
    sys.exit(0 if "error" not in result else 1)


if __name__ == "__main__":
    main()
